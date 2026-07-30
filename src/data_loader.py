"""
data_loader.py
--------------
Reads the 'Outputs' sheet of the Vista Energy integrated financial model and
turns it into tidy, chart-ready pandas objects.

Design notes
------------
* The Excel sheet is a *human-readable* report: sections stacked vertically,
  metric labels in column B, units in column C, and one column per fiscal year.
  Instead of hard-coding cell addresses (which break the moment a row is
  inserted), this module anchors on **text labels** and on the year header rows
  (e.g. "2022A", "2026E"). That makes the loader resilient to model edits.
* Every public function returns plain pandas objects so the Streamlit layer
  stays free of parsing logic.
* Nothing here imports Streamlit — the module can be unit-tested or reused in a
  notebook.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

#: Preferred location of the model, relative to the repository root.
DEFAULT_MODEL_PATH = Path(__file__).resolve().parents[1] / "data" / "vista_financial_model.xlsx"

#: Folders searched, in order, when the preferred file is not present.
SEARCH_DIRS = [
    Path(__file__).resolve().parents[1] / "data",
    Path(__file__).resolve().parents[1],
]


def resolve_model_path(path: str | Path | None = None) -> Path:
    """
    Locate the model workbook.

    Order of preference:
      1. An explicit path passed by the caller (or typed in the sidebar).
      2. ``data/vista_financial_model.xlsx``.
      3. The first ``.xlsx`` found in ``data/``, then in the repository root.

    Step 3 exists so the app keeps working when the workbook is committed under
    its original name — which usually contains spaces and punctuation that are
    awkward to hard-code. Excel's temporary ``~$`` lock files are skipped.
    """
    if path:
        candidate = Path(path)
        if candidate.exists():
            return candidate
        raise FileNotFoundError(f"Model file not found at {candidate}.")

    if DEFAULT_MODEL_PATH.exists():
        return DEFAULT_MODEL_PATH

    for directory in SEARCH_DIRS:
        if not directory.is_dir():
            continue
        matches = sorted(
            item for item in directory.glob("*.xlsx") if not item.name.startswith("~$")
        )
        if matches:
            return matches[0]

    raise FileNotFoundError(
        "No .xlsx model found. Place the workbook in data/ or pass an explicit path."
    )

#: Name of the sheet holding the summarised model outputs.
OUTPUT_SHEET = "Outputs"

#: A fiscal-year header looks like "2024A" (actual) or "2027E" (estimate).
YEAR_PATTERN = re.compile(r"^(19|20)\d{2}[AE]$")

#: Section titles written in ALL CAPS in column B of the Outputs sheet.
SECTION_TITLES = {
    "OPERATIONAL METRICS",
    "INCOME STATEMENT",
    "CASH FLOW STATEMENT",
    "BALANCE SHEET SUMMARY",
    "VALUATION SUMMARY",
    "SENSITIVITY ANALYSIS",
    "DCF CALCULATION",
    "SENSITIVITY ENGINE (BRENT × WACC)",
}


# --------------------------------------------------------------------------- #
# Container
# --------------------------------------------------------------------------- #


@dataclass
class ModelData:
    """Everything the dashboard needs, parsed once and cached."""

    #: Tidy long-format table: section | metric | unit | year | period | value
    tidy: pd.DataFrame = field(default_factory=pd.DataFrame)
    #: Single-value valuation figures (share price, EV, price targets, ...).
    valuation: Dict[str, float] = field(default_factory=dict)
    #: DCF build-up (WACC, PV of FCF, terminal value, equity value, ...).
    dcf: Dict[str, float] = field(default_factory=dict)
    #: Sensitivity grid — index = Brent ($/bbl), columns = WACC (fraction).
    sensitivity: pd.DataFrame = field(default_factory=pd.DataFrame)
    #: Ordered fiscal-year labels, e.g. ["2022A", ..., "2030E"].
    years: List[str] = field(default_factory=list)
    #: Free-text footnotes found in the sheet (sources, methodology).
    notes: List[str] = field(default_factory=list)

    # -- convenience accessors ---------------------------------------------- #

    def series(self, metric: str) -> pd.DataFrame:
        """Return the rows of `tidy` for one metric, ordered by year."""
        out = self.tidy[self.tidy["metric"].str.lower() == metric.lower()]
        if out.empty:
            raise KeyError(f"Metric not found in the model: {metric!r}")
        return out.sort_values("year_num").reset_index(drop=True)

    def value(self, metric: str, year: str) -> float:
        """Return a single metric/year figure (e.g. 'Revenue', '2026E')."""
        rows = self.series(metric)
        hit = rows[rows["year"] == year]
        if hit.empty:
            raise KeyError(f"No value for {metric!r} in {year!r}")
        return float(hit["value"].iloc[0])

    @property
    def last_actual_year(self) -> str:
        """The most recent reported (non-forecast) fiscal year."""
        actuals = [y for y in self.years if y.endswith("A")]
        return actuals[-1]

    @property
    def first_forecast_year(self) -> str:
        forecasts = [y for y in self.years if y.endswith("E")]
        return forecasts[0]


# --------------------------------------------------------------------------- #
# Low-level helpers
# --------------------------------------------------------------------------- #


def _cell(ws: Worksheet, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _find_label_row(ws: Worksheet, text: str, col: int = 2) -> Optional[int]:
    """Return the first row whose column `col` starts with `text` (case-insensitive)."""
    needle = text.strip().lower()
    for row in range(1, ws.max_row + 1):
        value = _cell(ws, row, col)
        if isinstance(value, str) and value.strip().lower().startswith(needle):
            return row
    return None


def _year_columns(ws: Worksheet, row: int) -> Dict[int, str]:
    """Map column index -> fiscal-year label for a header row such as row 5."""
    mapping: Dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        value = _cell(ws, row, col)
        if isinstance(value, str) and YEAR_PATTERN.match(value.strip()):
            mapping[col] = value.strip()
    return mapping


def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


# --------------------------------------------------------------------------- #
# Section parsers
# --------------------------------------------------------------------------- #


def _parse_statements(ws: Worksheet) -> tuple[pd.DataFrame, List[str], List[str]]:
    """
    Walk the sheet top-to-bottom and collect every `metric x year` figure that
    sits under a year header row. Returns (tidy dataframe, year labels, notes).
    """
    records: List[dict] = []
    notes: List[str] = []
    year_map: Dict[int, str] = {}
    section = "GENERAL"
    all_years: List[str] = []

    for row in range(1, ws.max_row + 1):
        label = _cell(ws, row, 2)

        # 1. Section header (ALL CAPS, no unit next to it).
        if isinstance(label, str) and label.strip().upper() in SECTION_TITLES:
            section = label.strip().upper()
            # Stop before the blocks that need bespoke parsing.
            if section in {"SENSITIVITY ANALYSIS", "DCF CALCULATION"} or section.startswith(
                "SENSITIVITY ENGINE"
            ):
                break
            continue

        # 2. Year header row -> refresh the column map used by the rows below.
        header = _year_columns(ws, row)
        if header:
            year_map = header
            for year in header.values():
                if year not in all_years:
                    all_years.append(year)
            continue

        if not isinstance(label, str) or not year_map:
            continue

        unit = _cell(ws, row, 3)
        values = {col: _cell(ws, row, col) for col in year_map}

        # 3. A row with no numbers under the year columns is a footnote.
        if not any(_is_number(v) for v in values.values()):
            if len(label.strip()) > 40:  # heuristic: prose, not an empty metric row
                notes.append(label.strip())
            continue

        metric = label.strip()
        for col, year in year_map.items():
            raw = values[col]
            if not _is_number(raw):
                continue
            records.append(
                {
                    "section": section,
                    "metric": metric,
                    "unit": (unit or "").strip() if isinstance(unit, str) else "",
                    "year": year,
                    "year_num": int(year[:4]),
                    "period": "Actual" if year.endswith("A") else "Forecast",
                    "value": float(raw),
                }
            )

    tidy = pd.DataFrame.from_records(records)
    # The valuation block holds one-off figures parked in a single column; they
    # are not a time series, so they live in `ModelData.valuation` instead.
    if not tidy.empty:
        tidy = tidy[tidy["section"] != "VALUATION SUMMARY"].reset_index(drop=True)
    all_years.sort(key=lambda y: int(y[:4]))
    return tidy, all_years, notes


def _parse_valuation(ws: Worksheet) -> Dict[str, float]:
    """
    The VALUATION SUMMARY block holds single figures in one column (K).
    Values are keyed by their label, lower-cased and snake-cased.
    """
    start = _find_label_row(ws, "VALUATION SUMMARY")
    if start is None:
        return {}

    out: Dict[str, float] = {}
    for row in range(start + 1, start + 25):
        label = _cell(ws, row, 2)
        if not isinstance(label, str):
            continue
        if label.strip().upper() in SECTION_TITLES and row != start:
            break
        # Scan the row for the first numeric value to the right of the unit column.
        for col in range(4, ws.max_column + 1):
            raw = _cell(ws, row, col)
            if _is_number(raw):
                out[_slug(label)] = float(raw)
                break
    return out


def _parse_dcf(ws: Worksheet) -> Dict[str, float]:
    """Parse the DCF build-up block (WACC, PVs, terminal value, value/share)."""
    start = _find_label_row(ws, "DCF CALCULATION")
    if start is None:
        return {}

    out: Dict[str, float] = {}
    for row in range(start + 1, min(start + 25, ws.max_row + 1)):
        label = _cell(ws, row, 2)
        if not isinstance(label, str):
            continue
        for col in range(3, ws.max_column + 1):
            raw = _cell(ws, row, col)
            if _is_number(raw):
                out[_slug(label)] = float(raw)
                break
        # Assumption pairs written side by side on the same row, e.g.
        # "WACC | 13% | ... | Exit multiple (EV/EBITDA) | 4.5x". The value is the
        # next numeric cell to the right of the label, allowing for blank spacers.
        for col in range(3, ws.max_column + 1):
            side_label = _cell(ws, row, col)
            if not isinstance(side_label, str) or not side_label.strip():
                continue
            for look in range(col + 1, min(col + 4, ws.max_column + 1)):
                side_value = _cell(ws, row, look)
                if _is_number(side_value):
                    out[_slug(side_label)] = float(side_value)
                    break

    # Yearly unlevered FCF, useful for the valuation waterfall.
    ufcf_row = _find_label_row(ws, "Unlevered Free Cash Flow")
    if ufcf_row:
        header = _year_columns(ws, ufcf_row - 1)
        out["ufcf_by_year"] = {  # type: ignore[assignment]
            year: float(_cell(ws, ufcf_row, col))
            for col, year in header.items()
            if _is_number(_cell(ws, ufcf_row, col))
        }
    return out


def _parse_sensitivity(ws: Worksheet) -> pd.DataFrame:
    """
    Parse the Brent x WACC grid into a DataFrame:
        index   -> Brent price ($/bbl)
        columns -> WACC (as a fraction, e.g. 0.13)
        values  -> implied value per share ($)
    """
    start = _find_label_row(ws, "SENSITIVITY ANALYSIS")
    if start is None:
        return pd.DataFrame()

    # The corner cell of the grid is the first row below the title that has
    # numbers running across it.
    header_row = None
    for row in range(start + 1, start + 8):
        across = [_cell(ws, row, col) for col in range(3, 9)]
        if sum(_is_number(v) for v in across) >= 3:
            header_row = row
            break
    if header_row is None:
        return pd.DataFrame()

    wacc_cols = {
        col: float(_cell(ws, header_row, col))
        for col in range(3, ws.max_column + 1)
        if _is_number(_cell(ws, header_row, col))
    }

    data: Dict[float, Dict[float, float]] = {}
    for row in range(header_row + 1, header_row + 20):
        brent = _cell(ws, row, 2)
        if not _is_number(brent):
            break
        data[float(brent)] = {
            wacc: float(_cell(ws, row, col))
            for col, wacc in wacc_cols.items()
            if _is_number(_cell(ws, row, col))
        }

    grid = pd.DataFrame(data).T.sort_index()
    grid.index.name = "Brent ($/bbl)"
    grid.columns.name = "WACC"
    return grid


def _slug(label: str) -> str:
    """'DCF price target' -> 'dcf_price_target'."""
    clean = re.sub(r"[^\w\s]", " ", label.strip().lower())
    return re.sub(r"\s+", "_", clean).strip("_")


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #


def load_model(path: str | Path | None = None, sheet: str = OUTPUT_SHEET) -> ModelData:
    """
    Read the model workbook and return a fully parsed :class:`ModelData`.

    Parameters
    ----------
    path : str | Path | None
        Path to the .xlsx model. When omitted, :func:`resolve_model_path` looks
        for ``data/vista_financial_model.xlsx`` and then for any other workbook
        in ``data/`` or in the repository root.
    sheet : str
        Sheet name holding the summary outputs. Defaults to ``"Outputs"``.

    Notes
    -----
    ``data_only=True`` reads the *cached values* of the formulas, so the workbook
    must have been opened and saved by Excel/LibreOffice at least once after the
    last edit — otherwise formula cells come back empty.
    """
    path = resolve_model_path(path)
    workbook = load_workbook(path, data_only=True, read_only=False)
    if sheet not in workbook.sheetnames:
        raise KeyError(
            f"Sheet {sheet!r} not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )
    ws = workbook[sheet]

    tidy, years, notes = _parse_statements(ws)
    if tidy.empty:
        raise ValueError(
            "No numeric data parsed. Open the workbook in Excel, save it, and try again "
            "(formula results must be cached in the file)."
        )

    data = ModelData(
        tidy=tidy,
        valuation=_parse_valuation(ws),
        dcf=_parse_dcf(ws),
        sensitivity=_parse_sensitivity(ws),
        years=years,
        notes=notes,
    )
    workbook.close()
    return data


def yoy_growth(frame: pd.DataFrame) -> pd.DataFrame:
    """Add a `yoy` column (year-on-year % change) to a single-metric frame."""
    out = frame.sort_values("year_num").copy()
    out["yoy"] = out["value"].pct_change()
    return out


def cagr(start_value: float, end_value: float, periods: int) -> float:
    """Compound annual growth rate. `periods` = number of years between the two."""
    if start_value <= 0 or periods <= 0:
        return float("nan")
    return (end_value / start_value) ** (1 / periods) - 1
